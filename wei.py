import datetime 
import random
import time
import os
import openpyxl
from selenium import webdriver
if(not os.path.exists('database.xlsx')):
    time.sleep(5)
    from openpyxl import Workbook
    book = Workbook()
    sheet = book.active
    sheet['B27']='Day'
    sheet['C27']='Drink'
    sheet['D27']='Food'
    sheet['E27']='Cake'
    sheet['F27']='Date'
    sheet['G27']='Total'
    sheet['H27']='Item'
    sheet['I27']='Debit'
    sheet['J27']='Credit'
    sheet['K27']='Balance'
    book.save('database.xlsx')
else:
    book = openpyxl.load_workbook('database.xlsx')
    sheet = book.active
row=int(sheet.dimensions.split(':')[1][1:])+1
try:
    f = open("username.txt", "r")
    a=f.read()
    f = open("password.txt", "r")
    b=f.read()
    f.close()
except:
    print('please provide username and password')
try:
    driver = webdriver.Chrome('chromedriver')
    driver.get('https://squareup.com/login')    
    username = driver.find_element_by_xpath('/html/body/div[1]/div/div/section/div[1]/div[3]/form/div[1]/input')
    username.send_keys(a)
    password = driver.find_element_by_xpath('/html/body/div[1]/div/div/section/div[1]/div[3]/form/div[3]/input')
    password.send_keys(b)
    sign_in_button = driver.find_element_by_xpath('/html/body/div[1]/div/div/section/div[1]/div[3]/form/div[6]/button')
    sign_in_button.click()
    time.sleep(random.randint(5,9))
    driver.get('https://squareup.com/dashboard/sales/reports/category-sales')
    time.sleep(random.randint(8,10))
    #time.sleep(random.randint(8,10))
except:
    print('unable to login trying again')
    time.sleep(1)

date= driver.find_element_by_xpath('/html/body/div[2]/div/div/div[2]/div[2]/div/div[2]/div/div[1]/div[1]/div/div[1]/div[1]/div/div/button').text
print(date)
day_name= ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday','Sunday']
day = datetime.datetime.strptime(date, '%m/%d/%Y').weekday()
print(day_name[day])
col=2
sheet.cell(row,col).value=day_name[day]
sheet.cell(row,col+4).value=date
sheet.cell(row,col+6).value='Food'
sheet.cell(row+1,col+6).value='Cake'
sheet.cell(row+2,col+6).value='Drink'
sheet.cell(row+3,col+6).value='Buy 7-11'
book.save('database.xlsx')
i=0
category=''
itemsold=''
grossales='  '
total=0
drinktotal=0
drinkitemtotal=0
def writeinexcel(drinktotal,drinkitemtotal):
    if(category=='FOOD'):
        sheet.cell(row,col+2).value=int(itemsold.replace(',',''))
        sheet.cell(row,col+7).value=float(grossales[1:].replace(',',''))
        pass
    elif(category=='Bakery'):
        sheet.cell(row,col+3).value=int(itemsold.replace(',',''))
        sheet.cell(row+1,col+7).value=float(grossales[1:].replace(',',''))
        pass
    else:
        drinkitemtotal+=int(itemsold.replace(',',''))
        drinktotal+=float(grossales[1:].replace(',',''))
        pass
    book.save('database.xlsx')
    return drinkitemtotal,drinktotal

while True:
    try:
        i+=1
        time.sleep(random.randint(3,5))
        category= driver.find_element_by_xpath('/html/body/div[2]/div/div/div[2]/div[2]/div/div[2]/div/div[1]/div[2]/div[4]/div/div/div/div/div[1]/div[1]/div['+str(i)+']/div/div[1]/span').text                
        print(category)                         
        itemsold= driver.find_element_by_xpath('/html/body/div[2]/div/div/div[2]/div[2]/div/div[2]/div/div[1]/div[2]/div[4]/div/div/div/div/div[1]/div[1]/div['+str(i)+']/div/div[2]/span').text
        print(itemsold)
        grossales= driver.find_element_by_xpath('/html/body/div[2]/div/div/div[2]/div[2]/div/div[2]/div/div[1]/div[2]/div[4]/div/div/div/div/div[1]/div[1]/div['+str(i)+']/div/div[3]/span').text
        print(grossales)
        total+=float(grossales[1:].replace(',',''))
        drinkitemtotal,drinktotal=writeinexcel(drinktotal,drinkitemtotal)    
    except:
        break
print(total,drinktotal)
sheet.cell(row,col+5).value=total
sheet.cell(row+2,col+7).value=drinktotal
sheet.cell(row,col+1).value=drinkitemtotal
book.save('database.xlsx')
driver.quit()
